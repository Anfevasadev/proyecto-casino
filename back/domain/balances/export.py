# -------------------------------------------
# back/domain/balances/export.py
# Propósito:
#   - Funciones para exportar reportes a PDF y Excel
# -------------------------------------------

from typing import Dict, Any
from io import BytesIO
from datetime import datetime


def generar_pdf_reporte(report: Dict[str, Any]) -> bytes:
    """
    Genera un PDF del reporte consolidado.
    
    Args:
        report: Diccionario con los datos del reporte
        
    Returns:
        bytes: Contenido del PDF
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    except ImportError:
        raise ImportError(
            "reportlab no está instalado. Instálalo con: pip install reportlab"
        )
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilo para título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1a237e'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Estilo para subtítulos
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#283593'),
        spaceAfter=12
    )
    
    # Título del reporte
    tipo_reporte = report.get('tipo_reporte', 'consolidado').upper()
    elements.append(Paragraph(
        f"REPORTE {tipo_reporte} DE CASINO",
        title_style
    ))
    elements.append(Spacer(1, 0.2*inch))
    
    # Información general
    info_data = [
        ['Periodo:', f"{report['period_start']} al {report['period_end']}"],
        ['Generado:', report['generated_at']],
        ['Generado por:', report['generated_by']]
    ]
    
    # Agregar información de casino o casinos incluidos
    if 'casino_nombre' in report:
        info_data.insert(0, ['Casino:', report['casino_nombre']])
    elif 'casinos_included' in report and report['casinos_included']:
        casinos_str = ', '.join([c['casino_nombre'] for c in report['casinos_included'][:3]])
        if len(report['casinos_included']) > 3:
            casinos_str += f" (+{len(report['casinos_included'])-3} más)"
        info_data.insert(0, ['Casinos:', casinos_str])
    
    # Agregar información de filtros si existen
    if 'filters_applied' in report:
        filters = report['filters_applied']
        filters_list = []
        if filters.get('marca'):
            filters_list.append(f"Marca: {filters['marca']}")
        if filters.get('modelo'):
            filters_list.append(f"Modelo: {filters['modelo']}")
        if filters.get('casino_id'):
            filters_list.append(f"Casino ID: {filters['casino_id']}")
        
        if filters_list:
            info_data.append(['Filtros:', ', '.join(filters_list)])
    
    info_table = Table(info_data, colWidths=[1.5*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e3f2fd')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Resumen por máquinas (solo si existe machines_summary)
    if 'machines_summary' in report and report['machines_summary']:
        elements.append(Paragraph("DESGLOSE POR MÁQUINA", subtitle_style))
        
        machine_headers = [
            ['ID', 'Marca', 'Modelo', 'Serial', 'IN', 'OUT', 'JACKPOT', 'BILLETERO', 'UTILIDAD']
        ]
        
        machine_data = []
        for machine in report['machines_summary']:
            machine_data.append([
                str(machine['machine_id']),
                machine.get('machine_marca', 'N/A')[:10],
                machine.get('machine_modelo', 'N/A')[:10],
                machine.get('machine_serial', 'N/A')[:10],
                f"${machine['in_total']:,.2f}",
                f"${machine['out_total']:,.2f}",
                f"${machine['jackpot_total']:,.2f}",
                f"${machine['billetero_total']:,.2f}",
                f"${machine['utilidad']:,.2f}"
            ])
        
        machine_table = Table(
            machine_headers + machine_data,
            colWidths=[0.4*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.9*inch, 0.9*inch, 0.9*inch, 1*inch, 1*inch]
        )
        
        machine_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        
        elements.append(machine_table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Totales por categoría
    elements.append(Paragraph("TOTALES POR CATEGORÍA", subtitle_style))
    
    # Usar category_totals si existe, sino construir desde otros campos
    if 'category_totals' in report:
        totals = report['category_totals']
        totals_data = [
            ['Categoría', 'Total'],
            ['IN TOTAL', f"${totals['in_total']:,.2f}"],
            ['OUT TOTAL', f"${totals['out_total']:,.2f}"],
            ['JACKPOT TOTAL', f"${totals['jackpot_total']:,.2f}"],
            ['BILLETERO TOTAL', f"${totals['billetero_total']:,.2f}"],
            ['', ''],
            ['UTILIDAD FINAL', f"${totals['utilidad_final']:,.2f}"]
        ]
    else:
        # Para reportes de participación
        totals_data = [
            ['Categoría', 'Total'],
            ['IN TOTAL', f"${report.get('in_total', 0):,.2f}"],
            ['OUT TOTAL', f"${report.get('out_total', 0):,.2f}"],
            ['JACKPOT TOTAL', f"${report.get('jackpot_total', 0):,.2f}"],
            ['BILLETERO TOTAL', f"${report.get('billetero_total', 0):,.2f}"],
            ['', ''],
            ['UTILIDAD TOTAL', f"${report.get('utilidad_total', 0):,.2f}"]
        ]
        
        # Agregar información de participación si existe
        if 'porcentaje_participacion' in report:
            totals_data.extend([
                ['', ''],
                ['PORCENTAJE PARTICIPACIÓN', f"{report['porcentaje_participacion']}%"],
                ['VALOR PARTICIPACIÓN', f"${report.get('valor_participacion', 0):,.2f}"]
            ])
    
    totals_table = Table(totals_data, colWidths=[3*inch, 2*inch])
    totals_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#4caf50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        ('GRID', (0, -1), (-1, -1), 1, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(totals_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Estadísticas
    stats_data = [
        ['Total de Máquinas:', str(report.get('total_machines', 0))],
        ['Total de Contadores Procesados:', str(report.get('total_counters', report.get('machines_processed', 0)))]
    ]
    
    stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
    stats_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))
    
    elements.append(stats_table)
    
    # Construir PDF
    doc.build(elements)
    
    pdf_content = buffer.getvalue()
    buffer.close()
    
    return pdf_content


def generar_excel_reporte(report: Dict[str, Any]) -> bytes:
    """
    Genera un archivo Excel del reporte consolidado.
    
    Args:
        report: Diccionario con los datos del reporte
        
    Returns:
        bytes: Contenido del Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl no está instalado. Instálalo con: pip install openpyxl"
        )
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Consolidado"
    
    # Estilos
    title_font = Font(name='Arial', size=14, bold=True, color='1a237e')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1976d2', end_color='1976d2', fill_type='solid')
    total_fill = PatternFill(start_color='4caf50', end_color='4caf50', fill_type='solid')
    total_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    tipo_reporte = report.get('tipo_reporte', 'consolidado').upper()
    ws['A1'] = f'REPORTE {tipo_reporte} DE CASINO'
    ws['A1'].font = title_font
    ws.merge_cells('A1:I1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Información general
    row = 3
    
    # Agregar información de casino o casinos incluidos
    if 'casino_nombre' in report:
        ws[f'A{row}'] = 'Casino:'
        ws[f'B{row}'] = report['casino_nombre']
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    elif 'casinos_included' in report and report['casinos_included']:
        ws[f'A{row}'] = 'Casinos:'
        casinos_str = ', '.join([c['casino_nombre'] for c in report['casinos_included'][:3]])
        if len(report['casinos_included']) > 3:
            casinos_str += f" (+{len(report['casinos_included'])-3} más)"
        ws[f'B{row}'] = casinos_str
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    ws[f'A{row}'] = 'Periodo:'
    ws[f'B{row}'] = f"{report['period_start']} al {report['period_end']}"
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = 'Generado:'
    ws[f'B{row}'] = report['generated_at']
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = 'Generado por:'
    ws[f'B{row}'] = report['generated_by']
    ws[f'A{row}'].font = Font(bold=True)
    
    # Desglose por máquina (solo si existe)
    if 'machines_summary' in report and report['machines_summary']:
        row += 3
        ws[f'A{row}'] = 'DESGLOSE POR MÁQUINA'
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:I{row}')
        
        # Headers
        row += 1
        headers = ['ID', 'Marca', 'Modelo', 'Serial', 'IN', 'OUT', 'JACKPOT', 'BILLETERO', 'UTILIDAD']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos de máquinas
        for machine in report['machines_summary']:
            row += 1
            ws.cell(row=row, column=1, value=machine['machine_id'])
            ws.cell(row=row, column=2, value=machine.get('machine_marca', 'N/A'))
            ws.cell(row=row, column=3, value=machine.get('machine_modelo', 'N/A'))
            ws.cell(row=row, column=4, value=machine.get('machine_serial', 'N/A'))
            ws.cell(row=row, column=5, value=machine['in_total'])
            ws.cell(row=row, column=6, value=machine['out_total'])
            ws.cell(row=row, column=7, value=machine['jackpot_total'])
            ws.cell(row=row, column=8, value=machine['billetero_total'])
            ws.cell(row=row, column=9, value=machine['utilidad'])
            
            # Formato de moneda para columnas numéricas
            for col in range(5, 10):
                cell = ws.cell(row=row, column=col)
                cell.number_format = '"$"#,##0.00'
                cell.border = border
            
            # Bordes para otras columnas
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border
    
    # Totales por categoría
    row += 3
    ws[f'A{row}'] = 'TOTALES POR CATEGORÍA'
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    
    # Usar category_totals si existe, sino usar campos directos
    if 'category_totals' in report:
        totals = report['category_totals']
        categories = [
            ('IN TOTAL', totals['in_total']),
            ('OUT TOTAL', totals['out_total']),
            ('JACKPOT TOTAL', totals['jackpot_total']),
            ('BILLETERO TOTAL', totals['billetero_total'])
        ]
        
        for label, value in categories:
            row += 1
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value).number_format = '"$"#,##0.00'
        
        # Utilidad final
        row += 2
        ws.cell(row=row, column=1, value='UTILIDAD FINAL').font = total_font
        ws.cell(row=row, column=1).fill = total_fill
        ws.cell(row=row, column=2, value=totals['utilidad_final']).font = total_font
        ws.cell(row=row, column=2).fill = total_fill
        ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
    else:
        # Para reportes de participación
        categories = [
            ('IN TOTAL', report.get('in_total', 0)),
            ('OUT TOTAL', report.get('out_total', 0)),
            ('JACKPOT TOTAL', report.get('jackpot_total', 0)),
            ('BILLETERO TOTAL', report.get('billetero_total', 0))
        ]
        
        for label, value in categories:
            row += 1
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value).number_format = '"$"#,##0.00'
        
        # Utilidad total
        row += 2
        ws.cell(row=row, column=1, value='UTILIDAD TOTAL').font = total_font
        ws.cell(row=row, column=1).fill = total_fill
        ws.cell(row=row, column=2, value=report.get('utilidad_total', 0)).font = total_font
        ws.cell(row=row, column=2).fill = total_fill
        ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
        
        # Participación si existe
        if 'porcentaje_participacion' in report:
            row += 2
            ws.cell(row=row, column=1, value='PORCENTAJE PARTICIPACIÓN:').font = Font(bold=True, size=11)
            ws.cell(row=row, column=2, value=f"{report['porcentaje_participacion']}%")
            
            row += 1
            ws.cell(row=row, column=1, value='VALOR PARTICIPACIÓN:').font = Font(bold=True, size=11, color='FF0000')
            ws.cell(row=row, column=1).fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
            ws.cell(row=row, column=2, value=report.get('valor_participacion', 0)).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=2).font = Font(bold=True, size=11, color='FF0000')
            ws.cell(row=row, column=2).fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
    
    # Estadísticas
    row += 3
    ws.cell(row=row, column=1, value='Total de Máquinas:').font = Font(bold=True)
    ws.cell(row=row, column=2, value=report.get('total_machines', 0))
    
    row += 1
    ws.cell(row=row, column=1, value='Total de Contadores:').font = Font(bold=True)
    ws.cell(row=row, column=2, value=report.get('total_counters', report.get('machines_processed', 0)))
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    excel_content = buffer.getvalue()
    buffer.close()
    
    return excel_content
